import csv
import zipfile
from pypdf import PdfReader
from conftest import ZIPPED_RESOURCES
from openpyxl import load_workbook


def test_second_page_in_pdf_should_have_substing():
    with zipfile.ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('lorem.pdf') as file:

            reader = PdfReader(file)
            first_page_text = reader.pages[0].extract_text()

            assert 'Integer at ultrices lorem' in first_page_text


def test_third_row_sixth_cell_should_have_france():
    with zipfile.ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('lorem.csv') as file:

            content = file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            third_row = csvreader[2]

            assert third_row[5] == 'France'


def test_second_row_third_column_should_have_male():
    with zipfile.ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('lorem.xlsx') as file:
            workbook = load_workbook(file)
            sheet = workbook.active

            assert sheet.cell(row=2, column=3).value == 'Male'
